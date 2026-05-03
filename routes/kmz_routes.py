import math
import os
import tempfile
import uuid
from datetime import datetime

import pandas as pd
from flask import request, jsonify, send_file, current_app
from werkzeug.utils import secure_filename

from routes import kmz_bp
from services.excel_parser import (
    parse_excel,
    group_by_station,
    get_icon_config,
    build_placemark_description,
)
from services.kmz_generator import generate_tower_kmz
from services.kmz_coverage_generator import generate_coverage_kmz


def _allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']


def _save_uploaded_file(file_storage):
    unique_id = str(uuid.uuid4())
    original_name = secure_filename(file_storage.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_{unique_id}_{original_name}"
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    file_storage.save(filepath)
    return filepath


def _cleanup_file(filepath):
    try:
        if filepath and os.path.exists(filepath):
            os.remove(filepath)
    except OSError as e:
        current_app.logger.warning(f'Falha ao remover arquivo temporário {filepath}: {e}')


def _watts_to_dbm(watts):
    if watts is None or watts <= 0:
        return 0.0
    return round(10.0 * math.log10(watts * 1000.0), 2)


def _dataframe_to_sites(df):
    sites = []
    for _, row in df.iterrows():
        site = {
            'name': str(row.get('Torre Estação', '')),
            'lat': float(row['Latitude']),
            'lon': float(row['Longitude']),
            'latitude': float(row['Latitude']),
            'longitude': float(row['Longitude']),
            'technology': str(row.get('Tecnologia', '')),
            'operadora': str(row.get('Torre Estação', '')),
            'numero_estacao': str(row.get('Numero Estacao', '')),
            'endereco': str(row.get('EnderecoEstacao', '')),
            'uf': str(row.get('SiglaUf', '')),
            'designacao_emissao': str(row.get('DesignacaoEmissao', '')),
            'freq_tx_mhz': float(row['FreqTxMHz']) if not pd.isna(row.get('FreqTxMHz')) else None,
            'freq_rx_mhz': float(row['FreqRxMHz']) if not pd.isna(row.get('FreqRxMHz')) else None,
            'azimuth': int(row['Azimute']) if not pd.isna(row.get('Azimute')) else 0,
            'antenna_gain_dbi': float(row['GanhoAntena']) if not pd.isna(row.get('GanhoAntena')) else 0.0,
            'frente_costa_db': float(row['FrenteCostaAntena']) if not pd.isna(row.get('FrenteCostaAntena')) else None,
            'beam_width_deg': float(row['AnguloMeiaPotenciaAntena']) if not pd.isna(row.get('AnguloMeiaPotenciaAntena')) else None,
            'tilt': int(row['AnguloElevacao']) if not pd.isna(row.get('AnguloElevacao')) else 0,
            'polarizacao': str(row.get('Polarizacao', '')),
            'tx_height_m': float(row['AlturaAntena']) if not pd.isna(row.get('AlturaAntena')) else 0.0,
            'cod_equipamento': str(row.get('CodEquipamentoTransmissor', '')),
            'tx_power_dbm': _watts_to_dbm(float(row['PotenciaTransmissorWatts'])) if not pd.isna(row.get('PotenciaTransmissorWatts')) else 0.0,
            'tx_power_watts': float(row['PotenciaTransmissorWatts']) if not pd.isna(row.get('PotenciaTransmissorWatts')) else None,
            'frequency_mhz': float(row['FreqTxMHz']) if not pd.isna(row.get('FreqTxMHz')) else None,
            'placemark_description': None,
            'icon_config': None,
        }
        operadora = site['operadora']
        tecnologia = site['technology']
        site['icon_config'] = get_icon_config(operadora, tecnologia)
        site['placemark_description'] = build_placemark_description(row)
        sites.append(site)
    return sites


def _process_excel_file(file_storage):
    excel_path = _save_uploaded_file(file_storage)
    try:
        df = parse_excel(excel_path)
        stations = group_by_station(df)
        sites = _dataframe_to_sites(df)
        return df, stations, sites, excel_path
    except Exception:
        _cleanup_file(excel_path)
        raise


def _calculate_bounds(sites):
    lats = [s.get('lat', s.get('latitude', 0)) for s in sites]
    lons = [s.get('lon', s.get('longitude', 0)) for s in sites]
    valid = [(la, lo) for la, lo in zip(lats, lons) if la != 0 and lo != 0]
    if not valid:
        return {'north': 0, 'south': 0, 'east': 0, 'west': 0}
    valid_lats, valid_lons = zip(*valid)
    return {
        'north': max(valid_lats),
        'south': min(valid_lats),
        'east': max(valid_lons),
        'west': min(valid_lons),
    }


# ---------------------------------------------------------------------------
# POST /api/kmz/upload
# ---------------------------------------------------------------------------

@kmz_bp.route('/upload', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({
            'success': False,
            'error': 'no_file',
            'message': 'Nenhum arquivo foi enviado na requisição.'
        }), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'empty_filename',
                        'message': 'O nome do arquivo está vazio.'}), 400
    if not _allowed_file(file.filename):
        return jsonify({
            'success': False, 'error': 'invalid_file_type',
            'message': f'Extensões aceitas: {", ".join(current_app.config["ALLOWED_EXTENSIONS"])}'
        }), 400

    excel_path = None
    try:
        excel_path = _save_uploaded_file(file)
        df = parse_excel(excel_path)
        stations = group_by_station(df)
        sites_data = _dataframe_to_sites(df)

        current_app.logger.info('Planilha processada: %d sites, %d estações', len(sites_data), len(stations))
        return jsonify({
            'success': True,
            'message': 'Planilha processada com sucesso.',
            'data': {
                'total_sites': len(sites_data),
                'total_stations': len(stations),
                'sites': sites_data,
                'stations': stations,
                'columns_detected': list(df.columns),
                'filename': secure_filename(file.filename)
            }
        }), 200

    except FileNotFoundError as e:
        return jsonify({'success': False, 'error': 'file_not_found', 'message': str(e)}), 404
    except ValueError as e:
        return jsonify({'success': False, 'error': 'validation_error', 'message': str(e)}), 422
    except Exception as e:
        current_app.logger.error(f'Erro ao processar planilha: {e}', exc_info=True)
        return jsonify({
            'success': False, 'error': 'processing_error',
            'message': 'Erro ao processar a planilha. Verifique o formato do arquivo.'
        }), 500
    finally:
        _cleanup_file(excel_path)


# ---------------------------------------------------------------------------
# POST /api/kmz/generate  -- aceita multipart Excel OU JSON {sites}
# ---------------------------------------------------------------------------

@kmz_bp.route('/generate', methods=['POST'])
def generate_kmz():
    show_sectors = request.form.get('show_sectors', 'false').lower() == 'true' \
        if request.form else request.get_json(silent=True) is None

    if 'excel_file' in request.files and request.files['excel_file'].filename != '':
        file = request.files['excel_file']
        if not _allowed_file(file.filename):
            return jsonify({
                'success': False, 'error': 'invalid_file_type',
                'message': f'Extensões aceitas: {", ".join(current_app.config["ALLOWED_EXTENSIONS"])}'
            }), 400

        excel_path = None
        try:
            excel_path = _save_uploaded_file(file)
            df = parse_excel(excel_path)
            sites = _dataframe_to_sites(df)

            kmz_filename = f"torres_rf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.kmz"
            kmz_path = os.path.join(current_app.config['UPLOAD_FOLDER'], kmz_filename)
            generate_tower_kmz(data=sites, output_path=kmz_path, show_sectors=show_sectors)

            current_app.logger.info('KMZ gerado via Excel: %s', kmz_filename)
            return send_file(kmz_path,
                             mimetype='application/vnd.google-earth.kmz',
                             as_attachment=True,
                             download_name=kmz_filename)

        except ValueError as e:
            return jsonify({'success': False, 'error': 'generation_error', 'message': str(e)}), 422
        except Exception as e:
            current_app.logger.error(f'Erro ao gerar KMZ: {e}', exc_info=True)
            return jsonify({
                'success': False, 'error': 'kmz_generation_error',
                'message': 'Erro ao gerar o arquivo KMZ.'
            }), 500
        finally:
            _cleanup_file(excel_path)

    body = request.get_json(silent=True)
    if not body or 'sites' not in body:
        return jsonify({
            'success': False, 'error': 'missing_data',
            'message': 'Envie um arquivo Excel (multipart/form-data com campo "excel_file") ou '
                       'JSON com lista de "sites".'
        }), 400

    sites = body['sites']
    show_sectors = body.get('show_sectors', False)

    if not isinstance(sites, list) or len(sites) == 0:
        return jsonify({
            'success': False, 'error': 'invalid_sites',
            'message': 'A lista de sites está vazia ou é inválida.'
        }), 400

    try:
        kmz_filename = f"torres_rf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.kmz"
        kmz_path = os.path.join(current_app.config['UPLOAD_FOLDER'], kmz_filename)
        generate_tower_kmz(data=sites, output_path=kmz_path, show_sectors=show_sectors)

        current_app.logger.info('KMZ gerado via JSON: %s', kmz_filename)
        return send_file(kmz_path,
                         mimetype='application/vnd.google-earth.kmz',
                         as_attachment=True,
                         download_name=kmz_filename)

    except ValueError as e:
        return jsonify({'success': False, 'error': 'generation_error', 'message': str(e)}), 422
    except Exception as e:
        current_app.logger.error(f'Erro ao gerar KMZ: {e}', exc_info=True)
        return jsonify({
            'success': False, 'error': 'kmz_generation_error',
            'message': 'Erro ao gerar o arquivo KMZ.'
        }), 500


# ---------------------------------------------------------------------------
# POST /api/kmz/preview  -- aceita Excel OU JSON {sites}
# ---------------------------------------------------------------------------

@kmz_bp.route('/preview', methods=['POST'])
def preview_sites():
    if 'excel_file' in request.files and request.files['excel_file'].filename != '':
        file = request.files['excel_file']
        if not _allowed_file(file.filename):
            return jsonify({
                'success': False, 'error': 'invalid_file_type',
                'message': f'Extensões aceitas: {", ".join(current_app.config["ALLOWED_EXTENSIONS"])}'
            }), 400

        excel_path = None
        try:
            excel_path = _save_uploaded_file(file)
            df = parse_excel(excel_path)
            stations = group_by_station(df)
            sites = _dataframe_to_sites(df)

            technologies = sorted(set(s.get('technology', '?') for s in sites))
            operators = sorted(set(s.get('operadora', '?') for s in sites))

            stations_sample = []
            for st_id, info in list(stations.items())[:5]:
                stations_sample.append({
                    'id': st_id,
                    'lat': info['info'].get('Latitude', info['info'].get('latitude', 0)),
                    'lon': info['info'].get('Longitude', info['info'].get('longitude', 0)),
                    'sectors': len(info['sectors']),
                    'endereco': str(info['info'].get('EnderecoEstacao', info['info'].get('endereco', ''))),
                })

            return jsonify({
                'success': True,
                'data': {
                    'total_stations': len(stations),
                    'total_sectors': len(sites),
                    'operators': operators,
                    'technologies': technologies,
                    'stations_sample': stations_sample,
                    'bounds': _calculate_bounds(sites),
                }
            }), 200

        except ValueError as e:
            return jsonify({'success': False, 'error': 'validation_error', 'message': str(e)}), 422
        except Exception as e:
            current_app.logger.error(f'Erro no preview: {e}', exc_info=True)
            return jsonify({'success': False, 'error': 'preview_error',
                            'message': 'Erro ao gerar preview.'}), 500
        finally:
            _cleanup_file(excel_path)

    body = request.get_json(silent=True)
    if not body or 'sites' not in body:
        return jsonify({
            'success': False, 'error': 'missing_data',
            'message': 'Envie um arquivo Excel (multipart) ou JSON com "sites".'
        }), 400

    sites = body['sites']
    if not isinstance(sites, list) or len(sites) == 0:
        return jsonify({
            'success': False, 'error': 'invalid_sites',
            'message': 'A lista de sites está vazia ou é inválida.'
        }), 400

    technologies = sorted(set(s.get('technology', s.get('tecnologia', '?')) for s in sites))
    operators = sorted(set(s.get('operadora', s.get('name', '?')) for s in sites))

    return jsonify({
        'success': True,
        'data': {
            'total_stations': len(sites),
            'total_sectors': len(sites),
            'operators': operators,
            'technologies': technologies,
            'stations_sample': [
                {
                    'id': str(s.get('numero_estacao', s.get('name', ''))),
                    'name': s.get('name', s.get('operadora', '')),
                    'lat': s.get('lat', s.get('latitude', 0)),
                    'lon': s.get('lon', s.get('longitude', 0)),
                    'technology': s.get('technology', s.get('tecnologia', '')),
                }
                for s in sites[:5]
            ],
            'bounds': _calculate_bounds(sites),
        }
    }), 200


# ---------------------------------------------------------------------------
# GET /api/kmz/template
# ---------------------------------------------------------------------------

TEMPLATE_COLUMNS = [
    "Torre Estação", "Numero Estacao", "EnderecoEstacao", "SiglaUf",
    "DesignacaoEmissao", "Tecnologia", "FreqTxMHz", "FreqRxMHz",
    "Azimute", "GanhoAntena", "FrenteCostaAntena", "AnguloMeiaPotenciaAntena",
    "AnguloElevacao", "Polarizacao", "AlturaAntena", "CodEquipamentoTransmissor",
    "PotenciaTransmissorWatts", "Latitude", "Longitude",
]

TEMPLATE_EXAMPLE = [
    ["CLARO S.A.", 1001, "Av. Paulista, 1000", "SP", "5M00G7W", "LTE", 2100.0, 1900.0, 0,
     18.0, 25.0, 65.0, 2, "X", 35.0, "TX-001", 20.0, -23.5505, -46.6333],
    ["CLARO S.A.", 1001, "Av. Paulista, 1000", "SP", "5M00G7W", "LTE", 2100.0, 1900.0, 120,
     18.0, 25.0, 65.0, 2, "X", 35.0, "TX-001", 20.0, -23.5505, -46.6333],
    ["CLARO S.A.", 1001, "Av. Paulista, 1000", "SP", "5M00G7W", "LTE", 2100.0, 1900.0, 240,
     18.0, 25.0, 65.0, 2, "X", 35.0, "TX-001", 20.0, -23.5505, -46.6333],
]


@kmz_bp.route('/template', methods=['GET'])
def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Template ANATEL"

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, row_data in enumerate(TEMPLATE_EXAMPLE, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or ''))
                       for r in range(1, len(TEMPLATE_EXAMPLE) + 2))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:S{len(TEMPLATE_EXAMPLE) + 1}"

    ws2 = wb.create_sheet("Instruções")
    instructions = [
        ["INSTRUÇÕES DE PREENCHIMENTO - Planilha ANATEL"],
        [""],
        ["As 19 colunas são obrigatórias. Abaixo a descrição de cada campo:"],
        ["Torre Estação", "Nome da operadora (CLARO, TELEFONICA, VIVO, TIM, OI, etc.)"],
        ["Numero Estacao", "Código numérico único da estação (ex: 1001)"],
        ["EnderecoEstacao", "Endereço completo da torre"],
        ["SiglaUf", "Sigla da Unidade Federativa (ex: SP, RJ, MG)"],
        ["DesignacaoEmissao", "Designação de emissão (ex: 5M00G7W)"],
        ["Tecnologia", "GSM, WCDMA, LTE ou NR"],
        ["FreqTxMHz", "Frequência de transmissão em MHz (ex: 2100)"],
        ["FreqRxMHz", "Frequência de recepção em MHz (ex: 1900)"],
        ["Azimute", "Ângulo de apontamento do setor em graus (0=Norte, 90=Leste)"],
        ["GanhoAntena", "Ganho da antena em dBi (ex: 18)"],
        ["FrenteCostaAntena", "Relação frente-costa em dB (ex: 25)"],
        ["AnguloMeiaPotenciaAntena", "Ângulo de meia potência em graus (ex: 65)"],
        ["AnguloElevacao", "Ângulo de elevação/tilt em graus (ex: 2)"],
        ["Polarizacao", "Tipo de polarização (X, V, H)"],
        ["AlturaAntena", "Altura da antena em metros (ex: 35)"],
        ["CodEquipamentoTransmissor", "Código do equipamento TX (ex: TX-001)"],
        ["PotenciaTransmissorWatts", "Potência do transmissor em Watts (ex: 20)"],
        ["Latitude", "Latitude em graus decimais (ex: -23.5505)"],
        ["Longitude", "Longitude em graus decimais (ex: -46.6333)"],
    ]
    for r, row in enumerate(instructions, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True, size=12, color="003366")

    buf = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(buf.name)
    buf.close()

    resp = send_file(buf.name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='template_anatel.xlsx')
    resp.call_on_close(lambda: _cleanup_file(buf.name))
    return resp


# ---------------------------------------------------------------------------
# POST /api/kmz/coverage
# ---------------------------------------------------------------------------

@kmz_bp.route('/coverage', methods=['POST'])
def generate_coverage():
    body = request.get_json(silent=True)

    if not body or 'stations' not in body:
        return jsonify({
            'success': False, 'error': 'missing_data',
            'message': 'O corpo da requisição deve conter uma lista de "stations".'
        }), 400

    stations = body['stations']
    show_labels = body.get('show_labels', True)
    show_signal_levels = body.get('show_signal_levels', True)
    show_sector_arrows = body.get('show_sector_arrows', False)

    if not isinstance(stations, list) or len(stations) == 0:
        return jsonify({
            'success': False, 'error': 'invalid_stations',
            'message': 'A lista de estações está vazia ou é inválida.'
        }), 400

    kmz_path = None
    try:
        kmz_filename = f"cobertura_rf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.kmz"
        kmz_path = os.path.join(current_app.config['UPLOAD_FOLDER'], kmz_filename)
        generate_coverage_kmz(
            stations_coverage=stations,
            output_path=kmz_path,
            show_labels=show_labels,
            show_signal_levels=show_signal_levels,
            show_sector_arrows=show_sector_arrows,
        )

        current_app.logger.info('KMZ de cobertura gerado: %s', kmz_filename)
        return send_file(kmz_pathf,
                         mimetype='application/vnd.google-earth.kmz',
                         as_attachment=True,
                         download_name=kmz_filename)

    except ValueError as e:
        return jsonify({'success': False, 'error': 'coverage_error', 'message': str(e)}), 422
    except Exception as e:
        current_app.logger.error(f'Erro ao gerar KMZ de cobertura: {e}', exc_info=True)
        return jsonify({
            'success': False, 'error': 'coverage_generation_error',
            'message': 'Erro ao gerar o KMZ de cobertura RF.'
        }), 500
